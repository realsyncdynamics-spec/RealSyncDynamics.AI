#!/usr/bin/env python3
"""
Build docs/economics/browser-fleet-economics.xlsx.

Idempotenter Builder fuer das Browser-Fleet-Oekonomie-Modell (Workstream 3,
Conversation 18.05.2026). Modelliert die Variable-Cost-Struktur des
Playwright-Scanner-Pools auf srv1622293 gegen die 5-Tier-Pricing-Struktur
aus src/config/pricing.ts.

Run:
    python3 scripts/build_browser_fleet_economics.py

Output:
    docs/economics/browser-fleet-economics.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ── Styling ────────────────────────────────────────────────────────────────

OBSIDIAN = "FF0A0A0B"
TITANIUM = "FFE2E2E2"
SECURITY_BLUE = "FF0052FF"
INPUT_YELLOW = "FFFFF4B8"
INPUT_BLUE = "FF0052FF"
HEADER_FILL = PatternFill("solid", fgColor=OBSIDIAN)
HEADER_FONT = Font(name="Arial", bold=True, color=TITANIUM, size=11)
INPUT_FILL = PatternFill("solid", fgColor=INPUT_YELLOW)
INPUT_FONT = Font(name="Arial", color=INPUT_BLUE, bold=True)
LABEL_FONT = Font(name="Arial", bold=True)
NORMAL_FONT = Font(name="Arial")
NOTE_FONT = Font(name="Arial", italic=True, color="FF666666", size=9)
SECTION_FONT = Font(name="Arial", bold=True, size=13, color=OBSIDIAN)
THIN = Side(style="thin", color=TITANIUM)
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT = "0.0%"
EUR = "#,##0.00 \\€"
EUR0 = "#,##0 \\€"
INT = "#,##0"


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = ALL_BORDER


def style_input(cell):
    cell.fill = INPUT_FILL
    cell.font = INPUT_FONT
    cell.border = ALL_BORDER


def style_label(cell):
    cell.font = LABEL_FONT
    cell.border = ALL_BORDER


def style_calc(cell):
    cell.font = NORMAL_FONT
    cell.border = ALL_BORDER


# ── Sheet 1: Assumptions ──────────────────────────────────────────────────


def build_assumptions(ws):
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50

    ws["A1"] = "Browser-Fleet Economics — Assumptions"
    ws["A1"].font = SECTION_FONT
    ws["A2"] = "Gelb hinterlegt = Input. Blau geschrieben. Aendere eine Zahl, alles Folgende rechnet sich neu."
    ws["A2"].font = NOTE_FONT

    rows = [
        ("INFRASTRUKTUR", None, None),
        ("VPS-Kosten srv1622293 (EUR/Monat)", 60, "Hostinger KVM-Linie. Bei Plan-Wechsel anpassen."),
        ("vCPU srv1622293", 4, "Aktuelle Allokation. PM2 + Playwright + Ollama teilen sich."),
        ("RAM (GB)", 8, ""),
        ("Reserviert fuer Ollama (vCPU)", 1, "qwen3:4b laeuft auf 1 vCPU."),
        ("Verfuegbare vCPU fuer Scanner", "=B4-B6", "auto"),
        ("Parallele Worker pro vCPU", 2, "Playwright headless mit ~500 MB pro Instanz."),
        ("Max parallele Scanner", "=B7*B8", "auto"),
        ("SCAN-PROFILE", None, None),
        ("Median Scan-Dauer (Sekunden)", 45, "Inkl. Page-Load, Network-Inventar, Consent-Check."),
        ("Overhead-Faktor (Retries, Idle, Warm-Up)", 1.5, "1.5 = 50% Aufschlag. 1.0 = keine Reserve."),
        ("Effektive Sekunden pro Scan", "=B11*B12", "auto"),
        ("Sekunden pro Monat (30 Tage)", "=30*24*3600", "auto"),
        ("Max Scans/Monat auf srv1622293", "=B9*B14/B13", "auto: parallel_capacity * seconds/month / seconds_per_scan"),
        ("STRIPE & ZAHLUNG", None, None),
        ("Stripe-Gebuehr (Prozent)", 0.014, "1.4% EU-Karten."),
        ("Stripe-Gebuehr (Fixum, EUR)", 0.25, ""),
        ("MIX-ANNAHME", None, None),
        ("Free-Audit-Anteil", 0.60, "One-time scan. Konvertiert ~5% zu Starter."),
        ("Starter-Anteil", 0.25, "1 Domain, monatlicher Re-Scan."),
        ("Growth-Anteil", 0.10, "3 Domains, taegliches Monitoring."),
        ("Agency-Anteil", 0.04, "10 Kundenseiten, White-Label."),
        ("Enterprise-Anteil", 0.01, "Unlimited (Fair-Use Cap unten)."),
        ("Mix-Summe (muss = 100% sein)", "=SUM(B20:B24)", "auto"),
        ("ENTERPRISE FAIR-USE", None, None),
        ("Fair-Use Domain-Cap Enterprise", 30, "Cap fuer Scans-Modellierung. Vertraglich >, im Pricing 'unlimited'."),
    ]

    for i, (label, value, note) in enumerate(rows, start=4):
        row = i
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_b = ws.cell(row=row, column=2)
        cell_c = ws.cell(row=row, column=3, value=note or "")
        cell_c.font = NOTE_FONT
        cell_c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if value is None:
            cell_a.font = Font(name="Arial", bold=True, color=TITANIUM, size=11)
            cell_a.fill = HEADER_FILL
            cell_b.fill = HEADER_FILL
            cell_c.fill = HEADER_FILL
            continue

        cell_a.font = LABEL_FONT
        cell_a.border = ALL_BORDER

        if isinstance(value, str) and value.startswith("="):
            cell_b.value = value
            style_calc(cell_b)
        else:
            cell_b.value = value
            style_input(cell_b)

        cell_b.alignment = Alignment(horizontal="right")
        if isinstance(value, float) and 0 < value < 1:
            cell_b.number_format = PCT
        elif label.startswith(("VPS", "Stripe-Geb")):
            cell_b.number_format = EUR
        elif isinstance(value, str) and "B14" in value:
            cell_b.number_format = INT
        elif isinstance(value, str) and "B9*B14/B13" in value:
            cell_b.number_format = INT

    return {
        "vps_cost": "'1_Assumptions'!$B$5",
        "max_scans_month": "'1_Assumptions'!$B$15",
        "stripe_pct": "'1_Assumptions'!$B$17",
        "stripe_fix": "'1_Assumptions'!$B$18",
        "mix_free": "'1_Assumptions'!$B$20",
        "mix_starter": "'1_Assumptions'!$B$21",
        "mix_growth": "'1_Assumptions'!$B$22",
        "mix_agency": "'1_Assumptions'!$B$23",
        "mix_enterprise": "'1_Assumptions'!$B$24",
        "fairuse_enterprise": "'1_Assumptions'!$B$27",
    }


# ── Sheet 2: Tier Economics ───────────────────────────────────────────────


def build_tier_economics(ws, refs):
    for col, width in zip("ABCDEFGHI", [18, 14, 14, 18, 18, 18, 18, 14, 50]):
        ws.column_dimensions[col].width = width

    ws["A1"] = "Pro Tier: Preis, Scan-Volumen, variable Kosten, Marge"
    ws["A1"].font = SECTION_FONT

    headers = [
        "Tier",
        "Preis (EUR)",
        "Domains",
        "Scans / Monat",
        "Var. Kosten",
        "Stripe-Gebuehr",
        "Deckungsbeitrag",
        "Marge %",
        "Modell-Notiz",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        style_header(c)

    cost_per_scan = f"({refs['vps_cost']}/{refs['max_scans_month']})"

    tiers = [
        ("Free Audit",  0,    1,                  1,        "One-time. Wird nur 1x pro Lead gerechnet."),
        ("Starter",     79,   1,                  1,        "Monatlicher Re-Scan, 1 Domain."),
        ("Growth",      249,  3,                  90,       "3 Domains * 30 Tage Monitoring = 90 Scans."),
        ("Agency",      699,  10,                 300,      "10 Kundenseiten * 30 Tage = 300 Scans."),
        ("Enterprise",  1500, f"={refs['fairuse_enterprise']}", None, "Fair-Use Cap aus Assumptions. Preis-Floor 1500 (ADR-001)."),
    ]

    for i, (name, price, domains, scans, note) in enumerate(tiers, start=4):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=price)
        if isinstance(domains, str):
            ws.cell(row=i, column=3, value=domains)
        else:
            ws.cell(row=i, column=3, value=domains)
        if scans is None:
            ws.cell(row=i, column=4, value=f"=C{i}*30")
        else:
            ws.cell(row=i, column=4, value=scans)

        ws.cell(row=i, column=5, value=f"=D{i}*{cost_per_scan}")
        ws.cell(row=i, column=6, value=f"=IF(B{i}=0,0,B{i}*{refs['stripe_pct']}+{refs['stripe_fix']})")
        ws.cell(row=i, column=7, value=f"=B{i}-E{i}-F{i}")
        ws.cell(row=i, column=8, value=f"=IF(B{i}=0,0,G{i}/B{i})")
        ws.cell(row=i, column=9, value=note).font = NOTE_FONT

        for col in range(1, 9):
            cell = ws.cell(row=i, column=col)
            cell.border = ALL_BORDER
            if col == 1:
                cell.font = LABEL_FONT
            elif col == 2:
                cell.number_format = EUR0
                style_input(cell) if name != "Enterprise" else style_calc(cell)
                if name == "Enterprise":
                    cell.font = NORMAL_FONT
                    style_input(cell)
            elif col == 3:
                style_input(cell) if not isinstance(domains, str) else style_calc(cell)
                cell.number_format = INT
            elif col == 4:
                style_calc(cell) if scans is None else style_input(cell)
                cell.number_format = INT
            elif col == 5:
                style_calc(cell)
                cell.number_format = EUR
            elif col == 6:
                style_calc(cell)
                cell.number_format = EUR
            elif col == 7:
                style_calc(cell)
                cell.number_format = EUR
            elif col == 8:
                style_calc(cell)
                cell.number_format = PCT

    ws["A10"] = "Cost-per-scan-Formel: (VPS-Kosten / Max-Scans-pro-Monat) — siehe Sheet 1, Zellen B5 und B15."
    ws["A10"].font = NOTE_FONT

    return {
        "tier_first_row": 4,
        "tier_last_row": 8,
    }


# ── Sheet 3: Capacity & Break-Even ────────────────────────────────────────


def build_capacity_breakeven(ws, refs, tier_refs):
    for col, width in zip("ABC", [44, 16, 50]):
        ws.column_dimensions[col].width = width

    ws["A1"] = "Capacity & Break-Even auf srv1622293"
    ws["A1"].font = SECTION_FONT
    ws["A2"] = "Gewichtetes Scan-Profil aus Mix-Annahme (Sheet 1) gegen Infrastrukturkapazitaet."
    ws["A2"].font = NOTE_FONT

    rows = [
        ("MIX (aus Sheet 1)", None, None),
        ("Free-Anteil",       f"={refs['mix_free']}",        None),
        ("Starter-Anteil",    f"={refs['mix_starter']}",     None),
        ("Growth-Anteil",     f"={refs['mix_growth']}",      None),
        ("Agency-Anteil",     f"={refs['mix_agency']}",      None),
        ("Enterprise-Anteil", f"={refs['mix_enterprise']}",  None),
        ("Summe (muss 100% sein)", "=SUM(B5:B9)", "100% = sauber kalibriert."),
        ("GEWICHTETER SCAN-BEDARF", None, None),
        (
            "Avg Scans / Tenant / Monat",
            f"=B5*'2_Tier_Economics'!D4 + B6*'2_Tier_Economics'!D5 + B7*'2_Tier_Economics'!D6 + B8*'2_Tier_Economics'!D7 + B9*'2_Tier_Economics'!D8",
            "Erwartungswert ueber alle Tiers.",
        ),
        ("Max Scans / Monat (Infra-Cap)", f"={refs['max_scans_month']}", "Sheet 1, B15."),
        (
            "Max Tenants auf srv1622293 (theoretisch)",
            "=B12/B11",
            "Bei aktuellem Mix.",
        ),
        ("Sicherheitsabschlag (Spike-Reserve)", 0.7, "70% nutzbar lassen 30% Headroom fuer Spikes."),
        ("Max Tenants empfohlen", "=B13*B14", "auto"),
        ("BREAK-EVEN", None, None),
        ("Fixkosten / Monat (VPS + Domains + Supabase)", 200, "Aus ADR-001: < 200 EUR/Mo total."),
        (
            "Avg Deckungsbeitrag / Tenant",
            f"=B5*'2_Tier_Economics'!G4 + B6*'2_Tier_Economics'!G5 + B7*'2_Tier_Economics'!G6 + B8*'2_Tier_Economics'!G7 + B9*'2_Tier_Economics'!G8",
            "Gewichteter Deckungsbeitrag aus Sheet 2 Spalte G.",
        ),
        ("Break-Even Tenant-Anzahl", "=B17/B18", "= Fixkosten / Avg-DB."),
        ("HEADROOM", None, None),
        ("Headroom bis Migration (Tenants)", "=B15-B19", "Wieviele Tenants ueber Break-Even hinaus auf srv1622293 passen."),
    ]

    for i, (label, value, note) in enumerate(rows, start=4):
        row = i
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_b = ws.cell(row=row, column=2)
        cell_c = ws.cell(row=row, column=3, value=note or "")
        cell_c.font = NOTE_FONT
        cell_c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if value is None:
            cell_a.font = Font(name="Arial", bold=True, color=TITANIUM, size=11)
            cell_a.fill = HEADER_FILL
            cell_b.fill = HEADER_FILL
            cell_c.fill = HEADER_FILL
            continue

        cell_a.font = LABEL_FONT
        cell_a.border = ALL_BORDER

        if isinstance(value, str) and value.startswith("="):
            cell_b.value = value
            style_calc(cell_b)
        else:
            cell_b.value = value
            style_input(cell_b)

        cell_b.alignment = Alignment(horizontal="right")

        if label.startswith(("Free-", "Starter-", "Growth-", "Agency-", "Enterprise-", "Summe ", "Sicherheits")):
            cell_b.number_format = PCT
        elif label in ("Fixkosten / Monat (VPS + Domains + Supabase)", "Avg Deckungsbeitrag / Tenant"):
            cell_b.number_format = EUR
        elif label.startswith(("Max ", "Avg Scans", "Break-Even", "Headroom", "Sicherheits")):
            cell_b.number_format = INT
        else:
            cell_b.number_format = INT


# ── Sheet 4: Scaling Roadmap ──────────────────────────────────────────────


def build_scaling_roadmap(ws):
    for col, width in zip("ABCDEF", [16, 18, 24, 38, 14, 38]):
        ws.column_dimensions[col].width = width

    ws["A1"] = "Scaling Roadmap — 5 Phasen, ADR-001-Trigger"
    ws["A1"].font = SECTION_FONT

    headers = ["Phase", "Tenant-Range", "Infrastruktur", "Worker-Pool", "Mtl. Kosten", "Migrations-Trigger"]
    for i, h in enumerate(headers, start=1):
        style_header(ws.cell(row=3, column=i, value=h))

    rows = [
        ("Phase 1 — Bootstrap", "0 – 50", "1 VPS (srv1622293)", "PM2 + systemd, 4–8 parallele Worker", "< 200 EUR", "—"),
        ("Phase 2 — Co-Pilot",  "50 – 200", "+ 2. VPS, Docker Compose", "Round-Robin via Traefik", "< 500 EUR", "Trigger D: > 8 parallele Worker sustained"),
        ("Phase 3 — NATS-Cutover", "200 – 500", "+ NATS JetStream", "Bus-getriebene Worker, kein LISTEN/NOTIFY mehr", "< 1.500 EUR", "Trigger A: > 5M Events/Tag oder p95 LISTEN > 500ms"),
        ("Phase 4 — Schema-per-Tenant", "500 – 1.000", "Postgres-Schema-Split, ggf. dedizierte DB pro Enterprise", "Dedizierte Worker-Pools fuer Enterprise", "< 4.000 EUR", "Trigger C: > 200 aktive Tenants oder p95-RLS > 300ms"),
        ("Phase 5 — Multi-Region EU", "1.000+", "Frankfurt + Paris/Stockholm, K8s (managed) oder Hetzner/OVH", "K8s-Operator fuer Worker-Lifecycle", "> 8.000 EUR", "Trigger D: > 4 Worker-VPS UND DevOps-Headcount"),
    ]

    for i, row in enumerate(rows, start=4):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = NORMAL_FONT
            c.border = ALL_BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if j == 1:
                c.font = LABEL_FONT
        ws.row_dimensions[i].height = 38

    ws["A10"] = "Trigger-Details: docs/adr/ADR-001-event-backbone.md §4."
    ws["A10"].font = NOTE_FONT


# ── Sheet 5: Recommendations ──────────────────────────────────────────────


def build_recommendations(ws):
    for col, width in zip("ABC", [6, 32, 80]):
        ws.column_dimensions[col].width = width

    ws["A1"] = "Operative Empfehlungen aus dem Modell"
    ws["A1"].font = SECTION_FONT

    for i, h in enumerate(["#", "Empfehlung", "Begruendung & Aktion"], start=1):
        style_header(ws.cell(row=3, column=i, value=h))

    recs = [
        (
            "Free-Tier hart limitieren",
            "Free-Audit ist Lead-Generator. Bei > 60% Mix-Anteil wird er zum Kostenfresser. Aktion: kein Account, kein Re-Scan, harte Rate-Limits pro IP/Domain.",
        ),
        (
            "Starter-Marge taeglich validieren",
            "EUR 79 mit 1 Domain + monatlichem Re-Scan ist enges Sweet-Spot. Stripe + Cost-per-Scan duerfen nicht > 25% des Tier-Preises kommen. Sheet 2 Zelle H5 wachhalten.",
        ),
        (
            "Growth ist das Profit-Center",
            "EUR 249 mit 3 Domains taeglich ist das Tier mit hoechster absoluter Marge. Sales-Fokus: Starter -> Growth-Upgrade nach 30 Tagen.",
        ),
        (
            "Agency-Quota an White-Label binden",
            "EUR 699 fuer 10 Kundenseiten. Wenn ein Agency-Kunde via API > 10 Domains pushed, automatische Quota-Warnung. Overage-Pricing oder Upgrade-Flow.",
        ),
        (
            "Enterprise: 'unlimited' braucht Fair-Use",
            "Vertraglich 'unlimitierte Domains' (pricing.ts). Praktisch begrenzt durch srv1622293-Kapazitaet. Fair-Use-Cap in Assumptions Sheet 1 B27 (Default 30 Domains) — bei Ueberschreitung Eskalation an Sales.",
        ),
        (
            "Zweiter Worker-VPS vor Tenant 30",
            "Headroom (Sheet 3 B21) wird bei aktuellem Mix bereits ab ~20-30 Tenants knapp. Phase 2 starten BEVOR sustained-overload eintritt.",
        ),
        (
            "Drift-Cron jitteren",
            "Bei taeglichem Monitoring fuer Growth + Agency: gestaffelte Slots ueber 24h, nicht alle um 00:00. Sonst Spike fuer 30-60 Min, Rest des Tages idle. Ohne Jitter: kuenstlich erzeugter Burst-Bedarf.",
        ),
        (
            "Stripe-SEPA fuer DE-Kunden anbieten",
            "Bei EUR 79 / 249 sind 1.4% + 0.25 EUR Karten-Fee schmerzhaft. SEPA Direct Debit: 0.8% + 0.25 EUR. Margen-Effekt auf Starter sichtbar in Sheet 2.",
        ),
    ]

    for i, (rec, why) in enumerate(recs, start=4):
        ws.cell(row=i, column=1, value=i - 3).font = LABEL_FONT
        ws.cell(row=i, column=1).border = ALL_BORDER
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=i, column=2, value=rec).font = LABEL_FONT
        ws.cell(row=i, column=2).border = ALL_BORDER
        ws.cell(row=i, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=i, column=3, value=why).font = NORMAL_FONT
        ws.cell(row=i, column=3).border = ALL_BORDER
        ws.cell(row=i, column=3).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[i].height = 56


# ── main ──────────────────────────────────────────────────────────────────


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("1_Assumptions")
    refs = build_assumptions(ws1)

    ws2 = wb.create_sheet("2_Tier_Economics")
    tier_refs = build_tier_economics(ws2, refs)

    ws3 = wb.create_sheet("3_Capacity_BreakEven")
    build_capacity_breakeven(ws3, refs, tier_refs)

    ws4 = wb.create_sheet("4_Scaling_Roadmap")
    build_scaling_roadmap(ws4)

    ws5 = wb.create_sheet("5_Recommendations")
    build_recommendations(ws5)

    out = Path(__file__).resolve().parent.parent / "docs" / "economics" / "browser-fleet-economics.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
